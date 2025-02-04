# Libs de manipulação para o Excel
import openpyxl as op
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side

# Libs variadas
import pandas as pd
from math import ceil

# Classe de estilos para o Excel
class Styles:
    def __init__(self):
        
        self.title_font = Font(name='Arial', bold=True, size=8) # Font para titúlos
        self.title_font_header = Font(name='Arial', bold=True, size=10) # Font para titúlos
        self.center_align = Alignment(vertical='center', horizontal='center') # Alinhamento no meio
        self.center_wrap_align = Alignment(vertical='center', horizontal='center', wrap_text=True)
        self.standard_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            bottom=Side(style='thin'),
            top=Side(style='thin')
        ) # Borda padrão
        self.standard_font = Font(name='Arial', size=8) # Fonte padrão
        self.standard_font_header = Font(name='Arial', size=10) # Fonte padrão
        self.center_left_align = Alignment(vertical='center', horizontal='left', wrap_text=True) # Fonte centralizada ao meio e esquerda
        self.top_rigth_align = Alignment(vertical='top', horizontal='right', wrap_text=True) # Alinhamento topo e direita

    # Método que aplica borda
    def apply_border(self, start, end, work_sheet):

        # Aplica um intervalo entre células
        for row in work_sheet[start:end]:
            # Para cada uma dentro desse intervalo aplique a borda
            for cell in row:
                # Tenta aplicar a borda para a célula
                try:
                    cell.border = self.standard_border
                
                # Retorna erro inesperado
                except Exception as err:
                    return {'status': False, 'error': f'Erro ao aplicar estilo de borda: "{str(err)}"'}        
    
    # Método para criar o cabeçalho do arquivo
    def create_header(self, worksheet):

        # Bloco superior esquerdo
        worksheet.merge_cells('A1:C4')
        
        # Bloco Technik
        worksheet.merge_cells('A5:C8')

        worksheet['A5'] = '=IMAGEM("https://technikgrupo.sharepoint.com/sites/transfer/Logos/Technik.png")' # Adicionando conteúdo de Imagem
        worksheet['A5'].alignment = self.center_align  # Alinhamento ao centro

        # Legenda Technik
        worksheet.merge_cells('A9:C10')

        worksheet['A9'] = 'CENTRO DE ENGENHARIA TECHNIK'
        worksheet['A9'].alignment = self.center_wrap_align # Alinhamento ao centro com quebra de linha
        worksheet['A9'].font = self.standard_font # Fonte padrão de texto

        # Título superior     
        worksheet.merge_cells('D1:G2')

        worksheet['D1'] = 'LISTA DE MATERIAIS'
        worksheet['D1'].alignment = self.center_align
        worksheet['D1'].font = self.title_font_header

        # Informações do cliente
        
        # Cliente
        worksheet['D3'] = 'CLIENTE'
        worksheet['D3'].alignment =  self.center_left_align
        worksheet['D3'].font = self.standard_font

        worksheet['D4'] = 'Client'
        worksheet['D4'].alignment =  self.center_left_align
        worksheet['D4'].font = self.standard_font

        worksheet.merge_cells('E3:L4')

        # Unidade
        worksheet['D5'] = 'UNIDADE'
        worksheet['D5'].alignment =  self.center_left_align
        worksheet['D5'].font = self.standard_font

        worksheet['D6'] = 'Plant'
        worksheet['D6'].alignment =  self.center_left_align
        worksheet['D6'].font = self.standard_font

        worksheet.merge_cells('E5:L6')

        # Área
        worksheet['D7'] = 'ÁREA'
        worksheet['D7'].alignment =  self.center_left_align
        worksheet['D7'].font = self.standard_font

        worksheet['D8'] = 'Area'
        worksheet['D8'].alignment =  self.center_left_align
        worksheet['D8'].font = self.standard_font

        worksheet.merge_cells('E7:L8')

        # Título
        worksheet['D9'] =  'Título'
        worksheet['D9'].alignment =  self.center_left_align
        worksheet['D9'].font = self.standard_font

        worksheet['D10'] =  'Title'
        worksheet['D10'].alignment =  self.center_left_align
        worksheet['D10'].font = self.standard_font

        worksheet.merge_cells('E9:l10')

        worksheet['E9'] = 'LISTA DE MATERIAIS DE TUBULAÇÃO'
        worksheet['E9'].alignment =  self.center_align
        worksheet['E9'].font = self.standard_font_header

        # Número do fluxograma
        worksheet.merge_cells('H1:H2')

        worksheet['H1'] =  'Nº'  # Label descritiva
        worksheet['H1'].alignment =  self.top_rigth_align
        worksheet['H1'].font = self.standard_font

        worksheet.merge_cells('I1:L2') # Campo para inserir valores
                
        row_index = 10

        return row_index
    
# Classe para o tratamento dos parafusos
class Screws:
    def __init__(self):

        # Dataframe utilizado para manipulação dos parafusos
        self.base_df= pd.DataFrame(
            {
                'class': [150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 150, 300, 300, 300, 300, 300, 300, 300, 300, 300, 300, 300, 300, 300, 300, 300, 300, 300, 300, 400, 400, 400, 400, 400, 400, 400, 400, 400, 400, 600, 600, 600, 600, 600],
                'diam_flange': ['1/2"', '3/4"', '1"', '1 1/4"', '1 1/2"', '2"', '2 1/2"', '3"', '4"', '6"', '8"', '10"', '12"', '14"', '16"', '18"', '20"', '24"', '1/2"', '3/4"', '1"', '1 1/4"', '1 1/2"', '2"', '2 1/2"', '3"', '4"', '6"', '8"', '10"', '12"', '14"', '16"', '18"', '20"', '24"', '4"', '6"', '8"', '10"', '12"', '14"', '16"', '18"', '20"', '24"', '1/2"', '3/4"', '1"', '1 1/4"', '1 1/2"'],
                'quantity': [4, 4, 4, 4, 4, 4, 4, 4, 8, 8, 8, 12, 12, 12, 16, 16, 20, 20, 4, 4, 4, 4, 4, 8, 8, 8, 8, 12, 12, 16, 16, 20, 20, 24, 24, 24, 8, 12, 12, 15, 16, 20, 20, 24, 24, 24, 4, 4, 4, 4, 4],
                'diam_screw': ['1/2"', '1/2"', '1/2"', '1/2"', '1/2"', '5/8"', '5/8"', '5/8"', '5/8"', '5/8"', '3/4"', '7/8"', '7/8"', '1"', '1"', '1 1/8"', '1 1/8"', '1 1/4"', '1/2"', '5/8"', '5/8"', '5/8"', '3/4"', '5/8"', '3/4"', '3/4"', '3/4"', '3/4"', '7/8"', '1"', '1 1/8"', '1 1/8"', '1 1/4"', '1 1/4"', '1 1/4"', '1 1/2"', '7/8"', '7/8"', '1"', '1 1/8"', '1 1/4"', '1 1/4"', '1 3/8"', '1 3/8"', '1 1/2"', '1 3/4"', '1/2"', '5/8"', '5/8"', '5/8"', '3/4"'],
                'length': ['2"', '2 1/4"', '2 1/4"', '2 1/2"', '2 1/2"', '2 3/4"', '3"', '3 1/4"', '3 1/4"', '3 1/2"', '3 5/6"', '4"', '4 1/4"', '4 1/2"', '4 3/4"', '5"', '5 1/2"', '6"', '2"', '2 1/4"', '2 1/2"', '2 3/4"', '3"', '3"', '3 1/2"', '3 3/4"', '4"', '4 1/4"', '4 3/4"', '5 1/2"', '6"', '6 1/4"', '6 1/2"', '6 3/4"', '7 1/4"', '8"', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
            }
        )

    # Método para obter todos os parafusos do dataframe
    def get_screws(self, dataframe):
        """
        Busca os valores de quantity, diam_screw e length no base_df,
        adicionando informações extras.
        """

        required_columns = ['Long Description (Family)', 'Spec', 'Size', 'Fixed Length']

        screws_df = dataframe.merge(
            self.base_df,
            left_on=['Pressure Class', 'Size'],
            right_on=['class', 'diam_flange'],
            how='left'
        ).drop(columns=['class', 'diam_flange'])  # Remove colunas duplicadas

        # Criar a coluna 'Long Description (Family)' usando vetorização
        screws_df['Long Description (Family)'] = 'PARAFUSO ' + screws_df['diam_screw'] + ' X ' + screws_df['length']

        # Filtra o tamanho de cada
        screws_df['Size'] = screws_df['diam_screw'] + 'x' + screws_df['length']

        # Multiplicar 'Fixed Length' pela quantidade
        screws_df['Fixed Length'] = screws_df['Fixed Length'] * screws_df['quantity']

        # Filtra o meu parafuso
        screws_df = screws_df[required_columns]

        ex = ExcelExtract()

        screws_df = ex.sum_unique(screws_df, required_columns)
        screws_df['Categorie'] = 'pç'

        return screws_df

# Classe para extração de dados do excel
class ExcelExtract:
    def __init__(self):
        pass
    
    # Método para formatar o valor do Fixed Length
    def ceil_format(self, value: float, categorie: str = 'm', extra_percent: float = 0):
        # Define value como new_value
        new_value = float(value)

        # Verifica se a categioria é metro
        if categorie == 'm':
            new_value /= 1000 # Altera a escala de centímetro para metro
        
        # Adiciona o valor extra de percentual
        new_value = ceil(new_value + (new_value * extra_percent))

        return float(new_value) # Retorna o valor em formato de float
    
    # Método para concatenar diversos DataFrames
    def concat(self, *args):
        """
        Concatena diversos DataFrames passados como argumentos. Se algum erro ocorrer, a exceção será levantada.

        Parameters:
            *args: DataFrames que devem ser concatenados.

        Returns:
            pd.DataFrame: DataFrame concatenado e ordenado.

        Raises:
            TypeError: Se algum dos argumentos não for do tipo DataFrame.
            Exception: Para erros inesperados durante a concatenação.
        """
        # Verifica se os argumentos foram passados
        if not args:
            raise ValueError('Nenhum dado foi passado para concatenar.')

        # Verifica se todos os argumentos são do tipo DataFrame
        for arg in args:
            if not isinstance(arg, pd.DataFrame):
                raise TypeError(f'O argumento "{arg}" não é do tipo DataFrame.')

        try:
            # Concatena os DataFrames
            mainDF = pd.concat(args, ignore_index=True)
            # Retorna os dados ordenados
            return mainDF.sort_values(by=['Long Description (Family)', 'Size', 'Spec'])

        except Exception as err:
            raise Exception(f'Erro inesperado: {str(err)}') from err

            
    # Método para somar valores únicos de cada dataframe
    def sum_unique(self, dataframe, required_columns):
        """
        Agrupa os dados do DataFrame por 'Long Description (Family)', 'Size', e 'Spec',
        e soma os valores da coluna 'Fixed Length'. Se ocorrer um erro, a exceção será levantada.
        
        Parameters:
            dataframe (pd.DataFrame): O DataFrame contendo os dados.
            
        Returns:
            pd.DataFrame: DataFrame agrupado e somado.

        Raises:
            ValueError: Se houver um erro entre os tipos de valores.
            KeyError: Se houver erro de chaves (colunas) no DataFrame.
            Exception: Para erros inesperados.
        """
        if 'Fixed Length' in required_columns:
            index = required_columns.index('Fixed Length')
            required_columns.pop(index)

        try:
            # Agrupa os dados do DataFrame que possuem dados semelhantes
            pipe_df = dataframe.groupby(required_columns, as_index=False)['Fixed Length'].sum()
            return pipe_df  # Retorna o DataFrame agrupado e somado

        # Tratamento de erros
        except ValueError as err:
            raise ValueError(f'Erro entre os tipos de valores: {str(err)}') from err

        except KeyError as err:
            raise KeyError(f'Erro de chaves da tabela: {str(err)}') from err

        except Exception as err:
            raise Exception(f'Erro inesperado: {str(err)}') from err
    
    # Método para somar valores únicos de cada dataframe
    def count_unique(self, dataframe, required_columns):
        """
        Agrupa os dados do DataFrame por 'Long Description (Family)', 'Size', e 'Spec',
        e conta as ocorrências para cada grupo. Se ocorrer um erro, a exceção será levantada.

        Parameters:
            dataframe (pd.DataFrame): O DataFrame contendo os dados.

        Returns:
            pd.DataFrame: DataFrame com a contagem de ocorrências agrupadas.

        Raises:
            ValueError: Se houver um erro entre os tipos de valores.
            KeyError: Se houver erro de chaves (colunas) no DataFrame.
            Exception: Para erros inesperados.
        """
        try:
            # Agrupa os dados e conta as ocorrências
            mainDF = dataframe.groupby(required_columns).size().reset_index(name='Fixed Length')
            return mainDF  # Retorna o DataFrame com a contagem

        # Tratamento de erros
        except ValueError as err:
            raise ValueError(f'Erro entre os tipos de valores: {str(err)}') from err

        except KeyError as err:
            raise KeyError(f'Erro de chaves da tabela: {str(err)}') from err

        except Exception as err:
            raise Exception(f'Erro inesperado: {str(err)}') from err
        
    def get_flange(self, dataframe):
        # Colunas requeridas para a planilha
        required_columns = ['Long Description (Family)', 'Spec', 'Size', 'Pressure Class']

        try:
            # Verifica se todas as colunas solicitadas existem
            missing_columns = [col for col in required_columns if col not in dataframe.columns]
            if missing_columns:
                raise KeyError(f"As seguintes colunas estão faltando: {', '.join(missing_columns)}")
            
            # Filtra as colunas
            flanges_df = dataframe[required_columns]
            flanges_df = flanges_df.dropna() # Deleta valores nulos

            flanges_df = self.count_unique(flanges_df, required_columns)  # Soma e filtra os valores únicos
            flanges_df['Categorie'] = 'pç'  # Define a unidade como metro para os dados
            flanges_df.drop(columns=['Pressure Class']) # Elimina a classe de pressão

            return flanges_df  # Retorna o DataFrame modificado

        except KeyError as err:
            raise KeyError(f'Não foi possível encontrar a chave "{str(err)}"') from err

        except Exception as err:
            raise Exception(f'Erro inesperado ocorreu: {str(err)}') from err
        
    def get_equipment(self, dataframe):
        # Colunas requeridas para a planilha
        required_columns = ['Long Description (Family)', 'Spec', 'Size']

        try:
            # Verifica se todas as colunas solicitadas existem
            missing_columns = [col for col in required_columns if col not in dataframe.columns]
            if missing_columns:
                raise KeyError(f"As seguintes colunas estão faltando: {', '.join(missing_columns)}")
            
            # Filtra as colunas
            equip_df = dataframe[required_columns]
            equip_df = equip_df.dropna() # Deleta valores nulos
            
            equip_df = equip_df[~equip_df['Long Description (Family)'].str.upper().str.contains(r'^(TUBO|PIPE|PARAF|BOLT|FLANGE)', regex=True)] # Filtra tirando as tubulações e as flanges

            equip_df = self.count_unique(equip_df, required_columns)  # Soma e filtra os valores únicos
            equip_df['Categorie'] = 'pç'  # Define a unidade como metro para os dados

            return equip_df  # Retorna o DataFrame modificado

        except KeyError as err:
            raise KeyError(f'Não foi possível encontrar a chave "{str(err)}"') from err

        except Exception as err:
            raise Exception(f'Erro inesperado ocorreu: {str(err)}') from err

    def get_pipe(self, dataframe):

        required_columns = ['Long Description (Family)', 'Spec', 'Size', 'Fixed Length']

        try:
            # Verifica se todas as colunas solicitadas existem
            missing_columns = [col for col in required_columns if col not in dataframe.columns]
            if missing_columns:
                raise KeyError(f"As seguintes colunas estão faltando: {', '.join(missing_columns)}")
            
            # Filtra as colunas
            pipeDF = dataframe[required_columns]
            pipeDF = pipeDF.dropna()

            pipeDF = self.sum_unique(pipeDF, required_columns)  # Soma e filtra os valores únicos
            pipeDF['Categorie'] = 'm'  # Define a unidade como metro para os dados

            return pipeDF  # Retorna o DataFrame modificado

        except KeyError as err:
            raise KeyError(f'Não foi possível encontrar a chave "{str(err)}"') from err

        except Exception as err:
            raise Exception(f'Erro inesperado ocorreu: {str(err)}') from err

    # Método para ler todos os arquivos
    def read_all_files(self, files, sheet_name):
        """Lê múltiplos arquivos Excel e concatena os dados em um único DataFrame."""

        # Validação dos parâmetros
        if not sheet_name or not files:
            raise ValueError('Os parâmetros "files" e "sheet_name" são necessários')

        main_df = pd.DataFrame()  # Cria um DataFrame vazio

        # Itera sobre a lista de arquivos
        for file in files:
            try: 
                temp_df = pd.read_excel(file, sheet_name=sheet_name)  # Lê a aba do Excel
                main_df = pd.concat([main_df, temp_df], ignore_index=True)  # Concatena os DataFrames
            
            # Tratamento de erros
            except FileNotFoundError:
                raise FileNotFoundError(f'O arquivo "{file}" não foi encontrado')

            except PermissionError:
                raise PermissionError(f'O arquivo "{file}" está aberto e não pode ser lido')

            except ValueError:
                continue
            
        # Filtrando dataframe
        try:
            main_df = main_df[main_df['Status'] == 'New'] # Filtra somente os dados novos]
            main_df['Spec'] = main_df['Spec'].astype(str).str.replace(r'(^SPEC|0)', '', regex=True).str.strip()

            # Retorna DataFrame
            return main_df

        # Tratamento de erros
        except KeyError as err:
            raise KeyError(f'A chave "{err}" não existe') from err
        
        except Exception as err:
            raise Exception(f'Erro inesperado ocorreu: {str(err)}') from err
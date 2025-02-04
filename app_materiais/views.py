from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
import json
import openpyxl as op
from io import BytesIO
import pandas as pd
from .classes import ExcelExtract, Screws, Styles

ex = ExcelExtract()
style = Styles()
screws = Screws()

def download_wb(request):
    if request.method == 'POST': # Verifica se o método é POST
        try: # Tenta executar o seguinte bloco:
            # Decodificar o corpo do request
            data = json.loads(request.body)

            # Criando workbook e worksheet
            wb = op.Workbook()
            ws = wb.active
            ws.title = "data"

            # Separa as SPECS
            specs = sorted({chave: list({d[chave] for d in data if chave in d}) for chave in {k for d in data for k in d}}.get('spec'))

            # Criando cabeçalho
            row_index =  style.create_header(ws)

            # Para cada Spec
            for i, spec in enumerate(specs):
                # Tentativa de execução do bloco
                try:
                    row_index += 1 # Incrementa uma linha para os dados
                    count_index = 0 # Reinicia a contagem de cada spec
                    
                    # Cria uma célula mesclada para adicionar o valor da Spec
                    ws.merge_cells(f'A{row_index}:L{row_index}')
                    ws[f'A{row_index}'] = f'SPEC {spec}'

                    # Estiliza a célula
                    ws[f'A{row_index}'].font = style.title_font
                    ws[f'A{row_index}'].alignment = style.center_align

                    ws.row_dimensions[row_index].height = 35 # Define o tamanho para 35

                    # Para cada linha da base de dados
                    for row in data:      
                        if row['spec'] == spec: # Verfiica se a Spec da linha é o mesmo da atual

                            row_index += 1 # Adiciona mais uma linha ao Excel
                            count_index += 1 # Adiciona mais uma contagem 

                            # Adiciona cada dado presente no dicionário
                            ws[f'A{row_index}'] = f'{i+1}.{count_index}'
                            ws[f'A{row_index}'].alignment = style.center_align
                            ws[f'A{row_index}'].font = style.standard_font

                            ws.merge_cells(f'B{row_index}:I{row_index}')
                            ws[f'B{row_index}'] = row['description']
                            ws[f'B{row_index}'].alignment = style.center_left_align
                            ws[f'B{row_index}'].font = style.standard_font

                            ws[f'J{row_index}'] = row['size']
                            ws[f'J{row_index}'].alignment = style.center_align 
                            ws[f'J{row_index}'].font = style.standard_font
                            
                            ws[f'K{row_index}'] = row['length']
                            ws[f'K{row_index}'].alignment = style.center_align 
                            ws[f'K{row_index}'].font = style.standard_font 

                            ws[f'L{row_index}'] = row['categorie']
                            ws[f'L{row_index}'].alignment = style.center_align
                            ws[f'L{row_index}'].font = style.standard_font

                            ws.row_dimensions[row_index].height = 35

                except Exception as err:
                    return render(request, 'error.html', {'error': f'Erro Inesperado: {str(err)}'})

            # Aplica borda nas células
            style.apply_border('A1', f'L{row_index}', ws)

            # Salvar planilha em memória
            output = BytesIO()
            wb.save(output)
            output.seek(0)


            # Configurar resposta
            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="Data.xlsx"'

            return response

        # Tratamento de erros
        except json.JSONDecodeError:
            return render(request, 'error.html', {'error': f'JSON Inválido'})
        # except Exception as err:
        #     return render(request, 'error.html', {'error': f'Erro Inesperado: {str(err)}'})
        
    # Erro de método
    return render(request, 'error.html', {'error': f'Utilização do método incorreto: {str(request.method)}'})

def handle_response(response: dict, request, error_template, data_key='data'):
    """
    View method for response validation and errors statements

    Args:
        response (dict): Dict response about status, data and errors
        request: Django HTTP Object
        error_template (str): Template to show error
        data_key (str): Dict key to access the data

    Returns:
        object: response data
        HttpResponse: Navigate to error template
    """
    try:
        if response['status']:
            return response[data_key]
        else:
            return render(request, error_template, {'error': response['error']})
    
    except KeyError as err:
        return render(request, 'error.html', {'error': f'KeyError encontrado: {str(err)}'})
        
    except Exception as err:
        return render(request, 'error.html', {'error': f'Erro Inesperado: {str(err)}'})

def upload_files(request):

    """
    View method for upload files in server to manipulation

    Args:
        request (str): Client request to server response

    Returns:
        HttpReponse: Server response
    """

    # Verify if request method is POST and files is not none
    if request.method == 'POST' and request.FILES.getlist('files'):
        try:
            files = request.FILES.getlist('files') # Declare files 
            
            # Filtra as colunas de tubulação
            pipe_df = ex.read_all_files(files, 'Pipe')
            pipe_df = ex.get_pipe(pipe_df)

            # Filtrando as flanges
            flanges_df = ex.read_all_files(files, 'Flange')
            blind_flanges_df = ex.read_all_files(files, 'Blind Flange')

            # Concatenando Flanges
            all_flanges_df = ex.concat(flanges_df, blind_flanges_df)
            
            # Obtendo flanges
            all_flanges_df = ex.get_flange(all_flanges_df)

            # Filtrando parafusos
            screws_df = screws.get_screws(all_flanges_df)

            # Filtrando os equipamentos
            equip_df = ex.read_all_files(files, 'Piping and Equipment')
            equip_df = ex.get_equipment(equip_df)

            # Concatenando os valores
            main_df = ex.concat(pipe_df, all_flanges_df, screws_df, equip_df)
            # Obtem o percentual adicional
            extra_percent = request.POST.get('percentual')
            extra_percent = float(extra_percent) / 100

            # Obtem o nome para o novo arquivo
            file_name = request.POST.get('file-name')

            # Cria o contexto para armazenar os dados
            context = []
            
            # Aplica a cada linha do DataFrame
            for _, row in main_df.iterrows():

                # Cria um dicionário para cada dado
                data = {
                    'description': row['Long Description (Family)'],
                    'spec': row['Spec'],
                    'size': row['Size'],
                    'length': ex.ceil_format(row['Fixed Length'], row['Categorie'], extra_percent),
                    'categorie': row['Categorie']
                }

                # Adiciona o dado ao contexto que será enviado
                context.append(data)
            
            # Retorne a tela de sucesso
            return render(request, 'success.html', {'data': context, 'filename': file_name})
    
        except Exception as e:
            return render(request, 'error.html', {'error': str(e)})
    else:
        return render(request, 'upload_files.html')
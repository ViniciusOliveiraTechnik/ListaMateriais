from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
import json
import openpyxl as op
from io import BytesIO
import pandas as pd
from .classes import ExcelExtract, Screws, Styles

ex = ExcelExtract()
style = Styles()

def download_wb(request):
    if request.method == 'POST': # Verifica se o método é POST
        try: # Tenta executar o seguinte bloco:

            # Decodificar o corpo do request
            data = json.loads(request.body)

            # Criando workbook e worksheet
            wb = op.Workbook()
            ws = wb.active
            ws.title = "data"

            # Separa as SPECS
            specs = sorted({chave: list({d[chave] for d in data if chave in d}) for chave in {k for d in data for k in d}}.get('spec'))

            row_index = 0

            for i, spec in enumerate(specs):

                row_index += 1
                count_index = 0
                
                ws.merge_cells(f'A{row_index}:L{row_index}')
                ws[f'A{row_index}'] = spec

                for row in data:      
                    if row['spec'] == spec:

                        row_index += 1
                        count_index +=1 

                        ws[f'A{row_index}'] = f'{i+1}.{count_index}'

                        ws.merge_cells(f'B{row_index}:I{row_index}')
                        ws[f'B{row_index}'] = row['description']

                        ws[f'J{row_index}'] = row['size']

                        ws[f'K{row_index}'] = row['length']

                        ws[f'L{row_index}'] = row['categorie']

            # Aplica borda nas células
            style.apply_border('A1', f'L{row_index}', ws)

            # Salvar planilha em memória
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Configurar resposta
            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="Data.xlsx"'

            return response

        # Tratamento de erros
        except json.JSONDecodeError:
            return render(request, 'error.html', {'error': f'JSON Inválido'})

        except Exception as err:
            return render(request, 'error.html', {'error': f'Erro Inesperado: {str(err)}'})
        
    # Erro de método
    return render(request, 'error.html', {'error': f'Utilização do método incorreto: {str(response.method)}'})

def handle_response(response: dict, request, error_template, data_key='data'):
    """
    View method for response validation and errors statements

    Args:
        response (dict): Dict response about status, data and errors
        request: Django HTTP Object
        error_template (str): Template to show error
        data_key (str): Dict key to access the data

    Returns:
        object: response data
        HttpResponse: Navigate to error template
    """
    try:
        if response['status']:
            return response[data_key]
        else:
            return render(request, error_template, {'error': response['error']})
    
    except KeyError as err:
        return render(request, 'error.html', {'error': f'KeyError encontrado: {str(err)}'})
        
    except Exception as err:
        return render(request, 'error.html', {'error': f'Erro Inesperado: {str(err)}'})

def upload_files(request):

    """
    View method for upload files in server to manipulation

    Args:
        request (str): Client request to server response

    Returns:
        HttpReponse: Server response
    """

    # Verify if request method is POST and files is not none
    if request.method == 'POST' and request.FILES.getlist('files'):
        files = request.FILES.getlist('files') # Declare files 
        
        # Pipe response
        response = ex.read_all_files(files, 'Pipe')
        pipeDF = handle_response(response, request, 'error.html')
        if isinstance(pipeDF, HttpResponse):
            return pipeDF
        
        # Equipment response
        response = ex.read_all_files(files, 'Piping and Equipment')
        equipDF = handle_response(response, request, 'error.html')
        if isinstance(equipDF, HttpResponse):
            return equipDF
        
        # Concatenando DataFrames
        response = ex.concat(pipeDF, equipDF)
        mainDF = handle_response(response, request, 'error.html')
        if isinstance(mainDF, HttpResponse):
            return mainDF
        
        # Obtem o percentual adicional
        extra_percent = request.POST.get('percentual')
        extra_percent = float(extra_percent) / 100

        # Obtem o nome para o novo arquivo
        file_name = request.POST.get('file-name')

        # Cria o contexto para armazenar os dados
        context = []
        
        # Aplica a cada linha do DataFrame
        for _, row in mainDF.iterrows():

            # Cria um dicionário para cada dado
            data = {
                'description': row['Long Description (Size)'],
                'spec': row['Spec'],
                'size': row['Size'],
                'length': ex.ceil_format(row['Fixed Length'], row['Categorie'], extra_percent),
                'categorie': row['Categorie']
            }

            # Adiciona o dado ao contexto que será enviado
            context.append(data)

        # Retorne a tela de sucesso
        return render(request, 'success.html', {'data': context, 'filename': file_name})

    else:
        return render(request, 'upload_files.html')